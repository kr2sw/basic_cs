"""
31: 고급 데이터 포맷 — csv, configparser, pickle, json
openpyxl(Excel)은 서드파티라 주석으로만 제공합니다.

# pip install openpyxl
from openpyxl import Workbook
wb = Workbook(); ws = wb.active
ws.append(["이름", "나이"]); ws.append(["홍길동", 30])
wb.save("users.xlsx")

from openpyxl import load_workbook
wb = load_workbook("users.xlsx"); ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
"""
import configparser
import csv
import io
import pickle


def csv_demo():
    print("=== 1) csv (메모리 StringIO 사용) ===")
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=["이름", "직업", "나이"])
    writer.writeheader()
    writer.writerow({"이름": "홍길동", "직업": "개발자", "나이": 30})
    writer.writerow({"이름": "김철수", "직업": "디자이너", "나이": 28})

    buffer.seek(0)
    reader = csv.DictReader(buffer)
    for row in reader:
        print(f"  {row['이름']} ({row['직업']}, {row['나이']}세)")
    print()


def config_demo():
    print("=== 2) configparser (메모리 StringIO 사용) ===")
    ini_text = """[database]
host = localhost
port = 5432

[logging]
level = DEBUG
"""
    config = configparser.ConfigParser()
    config.read_string(ini_text)
    print(f"  host={config['database']['host']}, port={config['database']['port']}")
    print(f"  level={config['logging']['level']}")

    # 수정 후 다시 문자열로 내보내기
    config["database"]["port"] = "3306"
    out = io.StringIO()
    config.write(out)
    print("  ==== 수정된 INI ====")
    print(out.getvalue().strip())
    print()


def pickle_demo():
    print("=== 3) pickle (바이트 직렬화) ===")
    data = {
        "name": "홍길동",
        "scores": [95, 88, 100],
        "meta": {"active": True, "tags": ("a", "b")},
    }
    blob = pickle.dumps(data)  # bytes
    print(f"  직렬화 크기: {len(blob)} bytes")
    restored = pickle.loads(blob)
    print(f"  복원 결과: {restored}")
    print(f"  값이 같고 타입 보존: {restored == data}, {type(restored['meta']['tags'])}")
    print()


def json_demo():
    print("=== 4) json (문자열 직렬화) ===")
    import json
    data = {"name": "홍길동", "scores": [95, 88, 100]}
    text = json.dumps(data, ensure_ascii=False, indent=2)
    print("  json 문자열:")
    for line in text.splitlines():
        print("   ", line)
    restored = json.loads(text)
    print(f"  복원: {restored}")
    print()


if __name__ == "__main__":
    csv_demo()
    config_demo()
    pickle_demo()
    json_demo()
